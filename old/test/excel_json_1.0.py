import json
import os
import re
import copy
import openpyxl
import pandas as pd


def is_cell_gray(cell):
    """
    终极版：无死角检测单元格底色是否为灰色 (支持 RGB, Indexed, 以及 Theme+Tint 计算)
    """
    if not getattr(cell, 'fill', None):
        return False

    colors_to_check = []
    if getattr(cell.fill, 'fgColor', None):
        colors_to_check.append(cell.fill.fgColor)
    if getattr(cell.fill, 'bgColor', None):
        colors_to_check.append(cell.fill.bgColor)
    if getattr(cell.fill, 'start_color', None):
        colors_to_check.append(cell.fill.start_color)

    for color in colors_to_check:
        c_type = getattr(color, 'type', None)

        # 1. 精确匹配 Theme + Tint (根据您的 Demo 探测结果)
        # Theme 0 (白色) 配合 -0.5 左右的 Tint (变暗50%) 视觉上就是 808080
        if c_type == 'theme':
            c_theme = getattr(color, 'theme', None)
            c_tint = getattr(color, 'tint', None)
            if c_theme == 0 and c_tint is not None:
                try:
                    tint_val = float(c_tint)
                    # 给一定的浮点数精度容差，通常 -0.45 到 -0.55 之间都是这块灰色
                    if -0.6 <= tint_val <= -0.4:
                        return True
                except (ValueError, TypeError):
                    pass
            continue # 如果是 theme 类型，RGB 属性通常会报错，直接跳过后面的 RGB 检查

        # 2. 匹配 RGB HEX 格式 (防御性取后缀)
        if c_type == 'rgb':
            try:
                rgb_val = str(color.rgb).upper()
                if rgb_val.endswith('808080') or rgb_val == '808080':
                    return True
            except Exception:
                pass

        # 3. 匹配 Excel 内置标准调色板 (索引 23 代表标准 50% 灰色 #808080)
        if c_type == 'indexed':
            if getattr(color, 'indexed', None) == 23:
                return True

    return False


def safe_convert_value(val):
    """安全地将数字转换为int/float，保持JSON整洁；若为Disable则视为空"""
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.lower() == "disable":
        return ""  # 遵循规则：Disable直接视为空，填充空值

    if isinstance(val, (int, float)):
        return val
    try:
        return float(val_str) if '.' in val_str else int(val_str)
    except ValueError:
        return val_str


def extract_data_with_openpyxl(file_path, sheet_name, possible_cols):
    """使用 openpyxl 提取带有样式截断逻辑的表格数据"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None, []

    sheet = wb[sheet_name]
    data = []
    header_row = -1
    cols_map = {}

    for r_idx, row in enumerate(sheet.iter_rows()):
        # 1. 寻找表头
        if header_row == -1:
            vals = [str(cell.value).strip() if cell.value else "" for cell in row]
            if 'Main Menu' in vals:
                header_row = r_idx
                for c_idx, cell in enumerate(row):
                    val = str(cell.value).strip() if cell.value else ""
                    if val in possible_cols or val in ['Range', 'Default', 'Ps.']:
                        cols_map[c_idx] = val
            continue

        # 2. 解析数据行，应用灰色截断规则
        row_dict = {}
        for c_idx, col_name in cols_map.items():
            cell = row[c_idx]

            # 【核心规则4】碰到灰色格子，当前行后续单元格全部忽略
            if is_cell_gray(cell):
                break

            val = cell.value
            row_dict[col_name] = val

        # 仅当有菜单层级数据时才保留该行
        if any(row_dict.get(col) for col in possible_cols):
            data.append(row_dict)

    wb.close()
    return data, list(cols_map.values())


def parse_sparse_tree(records: list, level_cols: list) -> dict:
    """带状态继承的树形解析器"""
    tree = {}
    current_path = [None] * len(level_cols)
    last_parent_path = []
    last_range = None

    for row in records:
        first_valid_idx = -1
        for i, col_name in enumerate(level_cols):
            val = row.get(col_name)
            if val is not None and str(val).strip() != "":
                first_valid_idx = i
                break

        if first_valid_idx == -1: continue

        for i in range(first_valid_idx, len(level_cols)):
            current_path[i] = None

        for i in range(first_valid_idx, len(level_cols)):
            val = row.get(level_cols[i])
            if val is not None and str(val).strip() != "":
                current_path[i] = str(val).strip()

        current_node = tree
        valid_path = []
        for i in range(len(level_cols)):
            node_name = current_path[i]
            if node_name is None: break
            valid_path.append(node_name)
            if node_name not in current_node:
                current_node[node_name] = {}
            current_node = current_node[node_name]

        # ======= 处理 Default, Range 与 Ps. =======
        default_val = row.get('Default')
        raw_range_val = row.get('Range')
        ps_val = row.get('Ps.')

        current_parent_path = valid_path[:-1]
        range_val = None

        # Range继承逻辑
        if raw_range_val is not None and str(raw_range_val).strip() != "":
            range_val = str(raw_range_val).strip()
            last_range = range_val
            last_parent_path = current_parent_path
        else:
            if current_parent_path == last_parent_path and last_range is not None:
                range_val = last_range
            else:
                last_range = None

        setting_path = valid_path.copy()
        if (range_val is None or str(range_val).strip() == ""):
            if len(setting_path) > 1:
                setting_path.pop()

        target_node = tree
        for p in setting_path:
            target_node = target_node[p]

        if default_val is not None and str(default_val).strip() != "":
            target_node["Default"] = safe_convert_value(default_val)

        if range_val:
            target_node["__Range__"] = range_val

        # 记录 Ps. 标记供后续使用
        if ps_val is not None and str(ps_val).strip() != "":
            target_node["__Ps__"] = str(ps_val).strip()

    return tree


def load_picture_mode_defaults(file_path: str) -> dict:
    """提取 Picture Mode 预设，分离 SDR 和 HDR 映射"""
    try:
        # 这里仅提取数据，无样式要求，故沿用 Pandas 保证速度
        excel_file = pd.ExcelFile(file_path)
        target_sheet = next((s for s in excel_file.sheet_names if "picture mode" in s.lower()), None)
        if not target_sheet:
            return {}

        df = pd.read_excel(file_path, sheet_name=target_sheet, header=None)

        mode_data = {"SDR": {}, "HDR": {}}
        current_context = None
        header_row = -1

        # 扫描并分类提取
        for idx, row in df.iterrows():
            vals = [str(v).strip() for v in row.values if pd.notna(v)]
            if not vals: continue

            # 定位模块标识
            if "SDR mode" in vals:
                current_context = "SDR"
                header_row = row
                for col_name in row.values[2:]:  # 假设前两列是 None 和 "SDR mode"
                    if pd.notna(col_name) and str(col_name).strip() != "":
                        mode_data["SDR"][str(col_name).strip()] = {}
                continue
            elif "HDR mode" in vals:
                current_context = "HDR"
                header_row = row
                for col_name in row.values[2:]:
                    if pd.notna(col_name) and str(col_name).strip() != "":
                        mode_data["HDR"][str(col_name).strip()] = {}
                continue

            # 填入预设值
            if current_context and header_row is not None:
                setting_name = str(row.iloc[1]).strip()  # 第2列通常是 Brightness, Contrast等
                if setting_name in ["", "nan", "None"]: continue
                clean_setting = setting_name.replace("(HUE)", "").strip()

                for i, col_val in enumerate(row.values[2:], start=2):
                    if pd.notna(col_val):
                        col_name = str(header_row.iloc[i]).strip()
                        if col_name in mode_data[current_context]:
                            mode_data[current_context][col_name][clean_setting] = safe_convert_value(col_val)

        return mode_data
    except Exception as e:
        print(f"    [-] Picture Mode 解析失败: {e}")
        return {}


def resolve_tree_references(tree, all_sheets_dict):
    """跨表合并链接器"""
    if not isinstance(tree, dict): return tree
    new_tree = {}
    for key, value in tree.items():
        if "Reference" in str(key):
            match = re.search(r'Reference\s+(.*)', str(key))
            if match:
                ref_sheet_name = match.group(1).strip()
                # 寻找目标表
                target_sheet = next((s for s in all_sheets_dict.keys() if ref_sheet_name.lower() in s.lower()), None)
                if target_sheet:
                    # 递归合并被引用表的内容
                    return resolve_tree_references(all_sheets_dict[target_sheet], all_sheets_dict)

        if key in ["Default", "__Range__", "__Ps__"]:
            new_tree[key] = value
        else:
            new_tree[key] = resolve_tree_references(value, all_sheets_dict)
    return new_tree


def apply_custom_business_rules(tree: dict, mode_defaults: dict) -> dict:
    if not isinstance(tree, dict): return tree

    # 1. 建立全局路径与节点索引 (用于后续的外链抓取)
    global_index = {}

    def build_index(n, path=[]):
        if not isinstance(n, dict): return
        for k, v in n.items():
            if isinstance(v, dict) and not str(k).startswith("__") and k != "Default":
                if k not in global_index:
                    global_index[k] = {"path": path, "node": v}
                build_index(v, path + [k])

    build_index(tree, ["Root"])

    def split_newline_keys(node):
        if not isinstance(node, dict): return
        keys_to_split = [k for k in list(node.keys()) if isinstance(k, str) and '\n' in k]
        for k in keys_to_split:
            val = node.pop(k)
            parts = [p.strip() for p in k.split('\n') if p.strip()]
            for part in parts:
                node[part] = copy.deepcopy(val)
        for k, v in node.items():
            if not str(k).startswith("__") and k != "Default":
                split_newline_keys(v)

    def parse_newline_ranges(node):
        if not isinstance(node, dict): return
        if "__Range__" in node:
            range_str = str(node["__Range__"])
            if '\n' in range_str:
                options = [o.strip() for o in range_str.split('\n') if o.strip()]
                for opt in options:
                    if opt not in node:
                        node[opt] = {}
        for k, v in node.items():
            if not str(k).startswith("__") and k != "Default":
                parse_newline_ranges(v)

    # 2. 组件分发、2D表默认值注入、外链抓取 三合一
    def propagate_picture_settings(node):
        if not isinstance(node, dict): return

        mode_keywords = {
            "Standard", "FPS", "MOBA", "RPG", "Racing", "Movie", "Reader",
            "sRGB", "Custom", "ECO", "HDR", "HDR Game", "HDR Movie",
            "HDR Vivid", "HDR Peak 1500"
        }
        child_keys = set(k for k in node.keys() if not str(k).startswith("__") and k != "Default")

        if len(child_keys.intersection(mode_keywords)) >= 2:
            richest_mode = None
            max_children = 0
            for mk in child_keys:
                if isinstance(node[mk], dict):
                    count = len([k for k in node[mk].keys() if not str(k).startswith("__") and k != "Default"])
                    if count > max_children:
                        max_children = count
                        richest_mode = mk

            target_defaults = None
            if "SDR" in mode_defaults and "Standard" in child_keys:
                target_defaults = mode_defaults["SDR"]
            elif "HDR" in mode_defaults and (
                    "HDR Game" in child_keys or "HDR Movie" in child_keys or "HDR Peak 1500" in child_keys):
                target_defaults = mode_defaults["HDR"]

            if richest_mode and max_children > 0:
                template = {}
                for k, v in node[richest_mode].items():
                    if not str(k).startswith("__") and k != "Default":
                        template[k] = copy.deepcopy(v)
                        if isinstance(template[k], dict) and "Default" in template[k]:
                            del template[k]["Default"]

                for mk in child_keys:
                    if not isinstance(node[mk], dict): node[mk] = {}

                    for tk, tv in template.items():
                        if tk not in node[mk]:
                            node[mk][tk] = copy.deepcopy(tv)

                    if target_defaults and mk in target_defaults:
                        for setting_name, def_val in target_defaults[mk].items():
                            if setting_name not in node[mk]:
                                if setting_name in global_index:
                                    node[mk][setting_name] = copy.deepcopy(global_index[setting_name]["node"])
                                    if "Default" in node[mk][setting_name]:
                                        del node[mk][setting_name]["Default"]

                                    clean_path = [p for p in global_index[setting_name]["path"] if
                                                  p not in ["Root", "Main Menu", "Setting", "Settings"]]
                                    node[mk][setting_name]["__Is_External_Link__"] = True
                                    node[mk][setting_name]["__Source_Path__"] = " -> ".join(clean_path + [setting_name])
                                else:
                                    node[mk][setting_name] = {}

                            node[mk][setting_name]["Default"] = def_val

        for v in node.values():
            if isinstance(v, dict):
                propagate_picture_settings(v)

    # 3. 【修改点：增加指令过滤器，只保留有实际意义的 Ps 标签】
    def restore_ps_keys(node):
        """扫描全树，将 __Ps__ 转换为规范的 Key-Value 格式存储，同时过滤掉内部映射指令"""
        if not isinstance(node, dict): return

        if "__Ps__" in node:
            ps_val = str(node["__Ps__"]).strip()

            # 核心过滤规则：剔除仅用于内部 2D 映射的指令词
            # 只有当 Ps 不是 "Picture Mode" 或 "HDR|Picture Mode" 时，才写入 JSON
            if ps_val not in ["Picture Mode", "HDR|Picture Mode"]:
                node["Ps"] = ps_val

        for v in node.values():
            if isinstance(v, dict):
                restore_ps_keys(v)

    def clean_temp_keys(node):
        """清理过程产生的临时标签"""
        if not isinstance(node, dict): return
        keys_to_delete = [k for k in node.keys() if
                          str(k).startswith("__") and k not in ["__Is_External_Link__", "__Source_Path__"]]
        for k in keys_to_delete:
            del node[k]

        for v in node.values():
            if isinstance(v, dict):
                clean_temp_keys(v)

    # 严格按照流水线顺序执行
    split_newline_keys(tree)
    parse_newline_ranges(tree)
    propagate_picture_settings(tree)
    restore_ps_keys(tree)  # 转化为 "Ps": "xxx"
    clean_temp_keys(tree)

    return tree

def build_super_menu(file_path: str, root_sheet: str, output_json: str = None):
    if not os.path.exists(file_path):
        print(f"[-] 文件不存在: {file_path}")
        return

    print(f"[*] 正在提取 Picture Mode 二维预设映射 (包含 Disable 过滤)...")
    mode_defaults = load_picture_mode_defaults(file_path)

    print(f"[*] 正在加载 Excel 并初始化带样式的全表索引...")
    all_sheets_data = {}

    # 获取所有的 Sheet 名称
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    possible_cols = ['Main Menu', 'Sub Menu\n Level 1', 'Sub Menu\n Level 2', 'Sub Menu\n Level 3',
                     'Sub Menu\nLevel 3+']

    for sheet_name in sheet_names:
        # 跳过 Picture Mode 表（已经单独解析过了）
        if "picture mode" in sheet_name.lower(): continue

        print(f"    -> 正在安全读取: '{sheet_name}'")
        records, actual_cols = extract_data_with_openpyxl(file_path, sheet_name, possible_cols)

        if records and actual_cols:
            all_sheets_data[sheet_name] = parse_sparse_tree(records, actual_cols)
            print(f"    + 已解析建树: '{sheet_name}'")

    if root_sheet not in all_sheets_data:
        print(f"[-] 错误: 根工作表 '{root_sheet}' 未找到。")
        return

    print(f"[*] 正在执行跨表组装 (Stitching)...")
    super_tree = resolve_tree_references(all_sheets_data[root_sheet], all_sheets_data)

    print(f"[*] 正在应用复杂的业务与映射规则...")
    final_tree = apply_custom_business_rules(super_tree, mode_defaults)

    json_output = json.dumps(final_tree, ensure_ascii=False, indent=4)
    if output_json:
        with open(output_json, 'w', encoding='utf-8') as f:
            f.write(json_output)
        print(f"[+] 超级菜单树 JSON 已生成，保存至: {output_json}")

    return final_tree


if __name__ == "__main__":
    # 请根据您的实际环境修改文件路径
    FILE_PATH = r"C:\Users\Administrator\Desktop\MO27Q28G.xlsx"
    ROOT_SHEET = "First Layer"
    OUTPUT_FILE = "../MO27Q28G.json"

    build_super_menu(FILE_PATH, ROOT_SHEET, OUTPUT_FILE)
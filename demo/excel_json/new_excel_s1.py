import os
import itertools
import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.views import SheetView


def is_gray_cell(cell):
    """
    [功能 2] 精准判断灰色：支持 RGB 808080 及主题色“白色背景1 深色50%”。
    """
    fill = cell.fill
    if not fill or getattr(fill, 'start_color', None) is None:
        return False
    color = fill.start_color
    # 检查 RGB
    if getattr(color, 'rgb', None) and isinstance(color.rgb, str):
        if '808080' in color.rgb.upper():
            return True
    # 检查主题色 (Theme 0/1, Tint ~ -0.5)
    if getattr(color, 'theme', None) is not None and getattr(color, 'tint', None) is not None:
        if color.theme in (0, 1) and -0.55 <= color.tint <= -0.45:
            return True
    return False


def process_osd_excel(input_path, sheet_names_str=None):
    """
    处理 OSD Excel 规格表
    [参数]：文件位置, 需要处理的工作表名称（分号分隔）
    """
    # [默认处理的工作表清单]
    if not sheet_names_str:
        sheet_names_str = "Control Button;Settings(Main Menu);AI Assistant(Main Menu);Game Assist(Main Menu);AI OLED Care Pro(Main Menu);System(Main Menu);Standby Menu;Picture Mode"

    target_sheets = [s.strip() for s in sheet_names_str.replace('；', ';').split(';')]

    print(f"[*] 正在加载文件: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    # 样式预设 [功能：加入对齐及表格线]
    thin_side = Side(border_style="thin", color="000000")
    std_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    std_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # [功能 1] 复制（保留）需要的工作表
    for sheet_name in wb.sheetnames:
        if sheet_name not in target_sheets:
            del wb[sheet_name]

    # [功能 4.1.1] 标准表头定义
    std_header = ["Main Menu", "Sub Menu\n Level 1", "Sub Menu\n Level 2", "Sub Menu\n Level 3",
                  "Sub Menu\nLevel 3+", "Range", "Default", "是否\n立即套用", "Function\nScale", "Remark"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"[*] 正在处理工作表: {sheet_name}")

        # [功能：平铺化处理] 解除合并并填充数据，解决 MergedCell 只读问题
        merged_ranges = list(ws.merged_cells.ranges)
        for m_range in merged_ranges:
            min_col, min_row, max_col, max_row = m_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(m_range))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if type(ws.cell(row=r, column=c)).__name__ == 'MergedCell':
                        ws._cells[(r, c)] = Cell(ws, row=r, column=c)
                    ws.cell(row=r, column=c).value = top_left_value

        # [功能 3] 去除所有图片
        ws._images = []

        # [功能 2] 当单元格为灰色时，删除整行
        rows_to_delete = []
        for row in ws.iter_rows():
            if any(is_gray_cell(c) for c in row):
                rows_to_delete.append(row[0].row)
        for r_idx in reversed(rows_to_delete):
            ws.delete_rows(r_idx)

        # [功能 4.1 / 4.1.4] Main Menu 与 Standby Menu 处理
        is_main_menu = "Main Menu" in sheet_name or "Standby Menu" in sheet_name
        header_row_idx, header_col_idx = -1, -1

        if is_main_menu:
            # [功能 4.1.2] 寻找表头行
            for row in ws.iter_rows(min_row=1, max_row=30):
                for cell in row:
                    val_str = str(cell.value) if cell.value else ""
                    if "Main Menu" in val_str:
                        header_row_idx, header_col_idx = cell.row, cell.column
                        break
                if header_row_idx != -1: break

            if header_row_idx != -1:
                # 写入标准表头
                for i, text in enumerate(std_header):
                    ws.cell(row=header_row_idx, column=header_col_idx + i).value = text

                # [功能 4.1.3] 区分主表与副表（寻找空行空列分隔）
                # 寻找底部空行
                bottom_row = ws.max_row
                for r in range(header_row_idx + 1, ws.max_row + 1):
                    if all(ws.cell(row=r, column=c).value is None for c in range(1, ws.max_column + 1)):
                        bottom_row = r - 1
                        break
                if bottom_row < ws.max_row:
                    ws.delete_rows(bottom_row + 1, ws.max_row - bottom_row)

                # 寻找右侧空列
                right_limit = header_col_idx + len(std_header)
                empty_col_idx = -1
                for c in range(right_limit, ws.max_column + 1):
                    if all(ws.cell(row=r, column=c).value is None for r in range(header_row_idx, ws.max_row + 1)):
                        empty_col_idx = c
                        break
                if empty_col_idx != -1:
                    ws.delete_cols(empty_col_idx, ws.max_column - empty_col_idx + 1)

                # 移至新文件（实际上是修剪当前表到 A1 开始）
                if header_row_idx > 1:
                    ws.delete_rows(1, header_row_idx - 1)
                    header_row_idx = 1
                if header_col_idx > 1:
                    ws.delete_cols(1, header_col_idx - 1)
                    header_col_idx = 1

        # [功能 5] 换行符嵌套展开（笛卡尔积）
        sub_menu_cols = []
        target_row = header_row_idx if header_row_idx != -1 else 1
        for cell in ws[target_row]:
            if cell.value and "Sub Menu" in str(cell.value):
                sub_menu_cols.append(cell.column)

        if sub_menu_cols:
            start_row = target_row + 1
            for r_idx in range(ws.max_row, start_row - 1, -1):
                row_data = {}
                needs_expand = False
                for col in sub_menu_cols:
                    val = ws.cell(row=r_idx, column=col).value
                    if val and '\n' in str(val):
                        needs_expand = True
                        row_data[col] = str(val).split('\n')
                    else:
                        row_data[col] = [val]

                if needs_expand:
                    combinations = list(itertools.product(*(row_data[c] for c in sub_menu_cols)))
                    ws.insert_rows(r_idx + 1, len(combinations) - 1)

                    # [修复点：全行向下填充数据]
                    orig_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]

                    for i, combo in enumerate(combinations):
                        curr_r = r_idx + i
                        # 先填充该行所有原始列的值（保证 Operation、Main Menu 不为空）
                        for c_idx, val in enumerate(orig_vals, 1):
                            ws.cell(row=curr_r, column=c_idx).value = val
                        # 覆盖 Sub Menu 展开的具体项
                        for j, col_idx in enumerate(sub_menu_cols):
                            ws.cell(row=curr_r, column=col_idx).value = combo[j]

        # [功能：视图修复与样式美化]
        ws.views.sheetView = [SheetView()]  # 修复 XML 视图报错
        # 去除顶部空行
        first_nonempty_row = 1
        for row in ws.iter_rows():
            if any(cell.value is not None for cell in row):
                first_nonempty_row = row[0].row
                break
        if first_nonempty_row > 1:
            ws.delete_rows(1, first_nonempty_row - 1)
        ws.freeze_panes = 'A2'
        for row in ws.iter_rows():
            for cell in row:
                cell.border = std_border
                cell.alignment = std_alignment

    # [功能] 移动到新的 xlsx 文件（保存为新文件名）
    output_file = input_path.replace(".xlsx", "_cleaned_flattened.xlsx")
    wb.save(output_file)
    print(f"[*] 处理成功！结果已保存至: {output_file}")
    return output_file


if __name__ == "__main__":
    # 使用示例
    process_osd_excel(r"C:\Users\Administrator\Desktop\Aorus OSD Tree_FO27Q24G_Q28G_20260401_tpv.xlsx")
    pass
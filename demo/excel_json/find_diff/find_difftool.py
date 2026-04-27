import openpyxl
from openpyxl.styles import PatternFill


def compare_excel(old_file_path, new_file_path, output_path):
    # 加载工作簿
    try:
        old_wb = openpyxl.load_workbook(old_file_path, data_only=True)
        new_wb = openpyxl.load_workbook(new_file_path)
    except Exception as e:
        print(f"读取文件失败: {e}")
        return

    # 定义颜色填充方案 (主题颜色)
    fill_modified = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色: 修改
    fill_added = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 浅绿: 新增
    fill_deleted = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红: 删除内容/位置

    # 获取所有同名的工作表
    common_sheets = [name for name in new_wb.sheetnames if name in old_wb.sheetnames]

    if not common_sheets:
        print("未找到同名工作表，请检查 Sheet 名称。")
        return

    for sheet_name in common_sheets:
        print(f"正在比对工作表: {sheet_name}...")
        new_ws = new_wb[sheet_name]
        old_ws = old_wb[sheet_name]

        # 确定比对的范围（取两表最大行列的并集）
        max_row = max(new_ws.max_row, old_ws.max_row)
        max_col = max(new_ws.max_column, old_ws.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                new_cell = new_ws.cell(row=row, column=col)
                old_cell = old_ws.cell(row=row, column=col)

                new_val = new_cell.value
                old_val = old_cell.value

                # 逻辑比对
                if new_val != old_val:
                    if old_val is None and new_val is not None:
                        # 新增
                        new_cell.fill = fill_added
                    elif old_val is not None and new_val is None:
                        # 删除 (原位置变空)
                        new_cell.fill = fill_deleted
                    else:
                        # 修改
                        new_cell.fill = fill_modified
                        # 可选：在批注里记录旧值
                        # new_cell.comment = openpyxl.comments.Comment(f"旧值: {old_val}", "System")

    # 保存结果
    try:
        new_wb.save(output_path)
        print(f"\n对比完成！结果已保存至: {output_path}")
        print("图例：[黄色] 修改 | [绿色] 新增 | [红色] 该位置原内容被删除")
    except Exception as e:
        print(f"保存失败: {e}")


# --- 使用示例 ---
if __name__ == "__main__":
    # 请替换为您实际的文件名
    OLD_EXCEL = r"C:\Users\Administrator\Desktop\Aorus OSD Tree_FO27Q24G_Q28G_20260401_tpv (1).xlsx"
    NEW_EXCEL = r"C:\Users\Administrator\Desktop\Aorus OSD Tree_FO27Q24G_Q28G_20260401_tpv.xlsx"
    RESULT_EXCEL = "diff_result.xlsx"

    compare_excel(OLD_EXCEL, NEW_EXCEL, RESULT_EXCEL)
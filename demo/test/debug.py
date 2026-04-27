import openpyxl


def inspect_excel_colors(file_path, sheet_name="Second Layer (Main menu)"):
    try:
        # data_only=True 获取公式计算后的值
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"文件读取失败: {e}")
        return

    # 模糊匹配工作表名称
    target_sheet = next((s for s in wb.sheetnames if sheet_name.lower() in s.lower()), None)

    if not target_sheet:
        print(f"[-] 找不到工作表，当前存在的工作表有: {wb.sheetnames}")
        return

    sheet = wb[target_sheet]
    print(f"[*] 正在扫描 '{target_sheet}' 的单元格颜色...")
    print("=" * 90)

    found_any_color = False

    # 遍历前 100 行，前 15 列（防止全表扫描过慢，覆盖菜单足够了）
    for row in sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=15):
        for cell in row:
            # 检查单元格是否有填充，且不是默认的透明(none)
            if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                color = cell.fill.fgColor

                # 提取底层的所有颜色元数据
                c_type = getattr(color, 'type', 'N/A')
                c_rgb = getattr(color, 'rgb', 'N/A')
                c_theme = getattr(color, 'theme', 'N/A')
                c_tint = getattr(color, 'tint', 'N/A')
                c_indexed = getattr(color, 'indexed', 'N/A')

                # 提取单元格内容（截取前15个字符方便查看）
                value_str = str(cell.value).replace('\n', ' ')[:15] if cell.value is not None else "[空]"

                # 打印日志
                print(f"坐标: {cell.coordinate:4} | "
                      f"内容: {value_str:<15} | "
                      f"类型: {str(c_type):<7} | "
                      f"RGB: {str(c_rgb):<10} | "
                      f"Theme: {str(c_theme):<4} | "
                      f"Tint: {str(c_tint)[:6]:<6} | "
                      f"Indexed: {str(c_indexed)}")

                # 如果代码发现了疑似灰色的RGB或索引，给予特殊标记
                if c_type == 'rgb' and str(c_rgb).upper().endswith('808080'):
                    print("  ---> [!] 系统判定：这是标准的RGB灰色！")

                found_any_color = True

    print("=" * 90)
    if not found_any_color:
        print("[-] 扫描完毕，没有在范围(1~100行)内检测到任何手动填充背景色的单元格。")
        print(
            "[-] 提示：如果表格里明明有灰色，说明它是通过【条件格式】或者【表格全局样式】生成的，openpyxl 无法直接读取这种动态颜色。")

    wb.close()


if __name__ == "__main__":
    # 请修改为您本机的测试文件路径
    FILE_PATH = r"C:\Users\Administrator\Desktop\MO27Q28G.xlsx"
    inspect_excel_colors(FILE_PATH)